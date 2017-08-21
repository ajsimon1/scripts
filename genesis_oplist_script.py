from openpyxl import *

# writing a quick script to pull out unique doctors list for compendium

# Source and destination filesnames
start_filename = r"C:\\Users\\adam\\Documents\\VitalAxis\\Genesis\\Genesis_OF.OP.xlsx"
dest_filename = r"C:\\Users\\adam\\Documents\\VitalAxis\\Genesis\\Genesis_ImplementationChecklist.xlsx"

# load spreadsheets in variables
wb_start = load_workbook(start_filename)
wb_dest = load_workbook(dest_filename)

# activate sheets
ws_start = wb_start.active
ws_dest = wb_dest['Workflow']

# define range
the_range = "L2:M53"

def iso_docs(a_range):
    count = 2
    final_list = []
    for row in ws_start.iter_rows(a_range):
        tup = (row[0].value, row[1].value)
        if tup in final_list:
            continue
        else:
            final_list.append(tup)
        count += 1
    r = 3
    for tup in final_list:
        ws_dest['B' + str(r)].value, ws_dest['C' + str(r)].value = tup
        r += 1
        wb_dest.save(dest_filename)

iso_docs(the_range)
