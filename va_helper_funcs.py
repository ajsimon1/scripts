# -*- coding: utf-8 -*-
"""
Created on Fri Jun 23 10:24:24 2017

@author: adam

helper functions used in time with VitalAxis to assist with various
Excel editing/reformatting needs
"""
from openpyxl import load_workbook

import itertools
"""
function for quantum path to compare the defauly GYN temps to their modified
version to see which temps in the system can be hidden from the backed
"""
# filenames
default_filename = "C:/Users/adam/Documents/VitalAxis/Quantum/GYNDxTeamplates_default.xlsx"
qp_adjusted_filename = "C:/Users/adam/Documents/VitalAxis/Quantum/GYN Dx Templates_VA default_with Site.xlsx"

# load workbooks
def_wb = load_workbook(default_filename)
qp_wb = load_workbook(qp_adjusted_filename)

# define ranges
def_range = "B2:C482"
qp_range = "C2:C127"


def dx_comp(wb_one, range_one, wb_two, range_two, default_filename):
    # load active worksheets
    ws_one = wb_one.active
    ws_two = wb_two.active
    for row_one in ws_one.iter_rows(range_one):
        for row_two in ws_two.iter_rows(range_two):
            if str(row_one[0].value).lower() == str(row_two[0].value).lower():
                row_one[1].value = ""
                break
            else:
                row_one[1].value = "Hide"
                continue

    wb_one.save(default_filename)

# dx_comp(def_wb, def_range, qp_wb, qp_range, default_filename)


def group(iterable, n=2):
    return zip(*([iter(iterable)] * n))


def grouper(iterable, n, fillvalue=None):
    "Collect data into fixed-length chunks or blocks"
    # grouper('ABCDEFG', 3, 'x') --> ABC DEF Gxx"
    args = [iter(iterable)] * n
    return itertools.zip_longest(*args, fillvalue=fillvalue)


def create_pairing_list(num_players, n=2):
    a_list = list(itertools.combinations(range(int(num_players) + 1)[1:], n))
    final_list = []
    master_list = []
    for item in a_list:

create_pairing_list(6)
