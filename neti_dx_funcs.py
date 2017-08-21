# -*- coding: utf-8 -*-
"""
Created on Thu May 25 16:08:36 2017

@author: adam
"""

# quick function to combine vertical cells into a single
# cell in excel
# UPDATE, couldnt get vertical combine function working

# new purpose is mapping the NETI site level derm dx codes

from openpyxl import load_workbook
import datetime

# start file
source_file = "NETI_siteLevelMapping_test.xlsx"

# icd file
icd_file = "Derm_ICD10mapping_test.xlsx"

# icd macro file
icd_macro_file = "Derm_ICD10mapping_test_macros.xlsm"

# safe icd macro file
safe_icd_macro_file = "Derm_ICD10mapping_test_macros.xlsx"

# load both workbooks
wb_source = load_workbook(source_file)
wb_icd = load_workbook(icd_file)
wb_icd_macro = load_workbook(icd_macro_file)
wb_safe_icd_macro = load_workbook(safe_icd_macro_file)

# load active tabs
ws_source = wb_source['actual_match']
ws_icd = wb_icd.active

# define both ranges
source_range = "B3:C96"
icd_range = "A2:F125216"
source_range_test = "B3:C5"
icd_range_test = "A2:F30000"

def neti_icd_mapping(wb_source, source_range, wb_icd, icd_range):
    ws_source = wb_source['actual_match']
    ws_icd = wb_icd.active
    print('Initializing...')
    for source_row in ws_source.iter_rows(source_range):
        print('Processing row {} of 93 at {}'.format(source_row[0].row, datetime.datetime.now()))
        for icd_row in ws_icd.iter_rows(icd_range):
            if str(source_row[0].value) == str(icd_row[0].value):
                icd_row[5].value = source_row[1].value
        wb_icd.save(safe_icd_macro_file)

add_extra_range = "A2:E125216"
add_absent_range = "A2:D125216"

def add_extra_site(wb, a_range):
    ws = wb.active
    for row in ws.iter_rows(a_range):
        if not row[0].value:
            row[0].value = ws[str(row[0].column) + str(int(row[0].row - 1))].value
            row[1].value = "Tarsal"
            row[2].value = ws[str(row[2].column) + str(int(row[2].row - 1))].value
            row[3].value = ws[str(row[3].column) + str(int(row[3].row - 1))].value
            row[4].value = ws[str(row[4].column) + str(int(row[4].row - 1))].value
    wb.save("Derm_ICD10mapping_test_macros.xlsx")
            

def fix_absent_left_right(wb, a_range):
    ws = wb.active
    check_list = ['Auricular', 'Deltoid', 'Digit', 'Fibula', 'Tibia', 'Tibial', 'Flank', 'Glabella', 'Gluteal Crease', 'Hallux']
    for row in ws.iter_rows(a_range):
        for item in check_list:
            if row[1].value == item:
                if ws[str(row[3].column) + str(int(row[3].row - 1))].value:
                    row[3].value = ws[str(row[3].column) + str(int(row[3].row - 1))].value
    wb.save("Derm_ICD10mapping_test_macros.xlsx")
        
neti_icd_mapping(wb_source,source_range,wb_safe_icd_macro,icd_range)
# add_extra_site(wb_icd_macro, add_extra_range)
# fix_absent_left_right(wb_safe_icd_macro, add_absent_range)

"""
total_range = 'B2:C10'

def vert_combine(a_range, ws):
    for row in ws.iter_rows(a_range):
        if row[0].value:
            continue
        else:
            # how to combine blank dx code with the previous
            # need a -1 to access last dx summary cell
            # can i iterate over the rows in reverse?
            # read iter_rows docs specifically rawCell
            # can get the coordinates during iteration, just need to
            # subtract and can concat to previous cell
            # question then becaomes how do you continue through iter and
            # append to first cell
            collected += row[1].value
            if ws[str(row[0].column+str(int(row[0].row + 1)))].value:
                ws[str(row[1].column+str(int(row[1].row - 1)))].value += '/n'row[1].value
            else:
            collected += row[1].value
            
    wb_source.save(start_file)

def vert_combine2(a_range, ws):
    for row in ws.iter_rows(a_range):
        if row[0].value:
            continue
        else:
            # how to combine blank dx code with the previous
            # need a -1 to access last dx summary cell
            # can i iterate over the rows in reverse?
            # read iter_rows docs specifically rawCell
            # can get the coordinates during iteration, just need to
            # subtract and can concat to previous cell
            # question then becaomes how do you continue through iter and
            # append to first cell
            if ws[str(row[0].column+str(int(row[0].row + 1)))].value:
                ws[str(row[1].column+str(int(row[1].row - 1)))].value += row[1].value
            else:
            count = 1
            collected = ''
            collected += row[1].value
            
    wb_source.save(start_file)
"""

