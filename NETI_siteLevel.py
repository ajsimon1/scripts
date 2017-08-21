import openpyxl

"""
set prefixes in tuple
iterate through dr olearys codes
look for match pre decimal prefix
if prefix in row[2] in tuple
copy row[0] add to template sheet row[0], comma seprated

slightly modified code on 8/10 for BCC templates
"""

# set filepaths in variables
leary = "ICD10_DX_NETI (1).xlsx"
template = "NETI_BCC.xlsx"
# load workbooks
leary_wb = openpyxl.load_workbook(leary)
template_wb = openpyxl.load_workbook(template)
# set ranges
leary_range = "A5:C163" #163
template_range = "A2:E293" #293

# start formula
def set_icd(wb_1, range_1, wb_2, range_2):
    # set constant
    PREFIXES = ('C44') # previously used 'D04', 'D23', 'D22', 'D03', 'C43'
    # add a 9 to C43 for all sites after the first 5 post decimal
    # set active sheets
    leary_ws = wb_1.active
    templates_ws = wb_2.active
    for row in leary_ws.iter_rows(range_1):
        try:
            split_prefix = row[2].value.split('.')[0]
            print(split_prefix)
        except AttributeError:
            row[2].value = 'NULL'
        # print(split_prefix)
        if split_prefix in PREFIXES:
            for row_a in templates_ws.iter_rows(range_2):
                split_postfix = row_a[4].value.split('.')[1]
                dx_list_split_postfix = row[2].value.split('.')[1]
                try:
                    final_postfix = split_postfix[0] + dx_list_split_postfix[1] + split_postfix[2]
                except IndexError:
                    final_postfix = split_postfix[0] + dx_list_split_postfix[1]
                row_a[0].value += ',' + row[0].value + '({}.{})'.format(split_prefix, final_postfix)
                # print(row_a[0].value)
    wb_2.save(template)

set_icd(leary_wb, leary_range, template_wb, template_range)
