from openpyxl import *

import warnings
import re

"""
taking spreadsheet sent by Amanada and re formatting into VA standard
"""

source_file = "Revised Sagis client list for Vitalaxis.as2ndedition.xlsx"
dest_file = "Sagis_OF.OP_test.xlsx"

total_range = "A2:X555"

#org_name_range = "B1:B749"

warnings.simplefilter("ignore")

wb_source = load_workbook(source_file)
wb_dest = load_workbook(dest_file)

ws_source = wb_source.active
ws_dest = wb_dest["Facilities"]
wsPhys_dest = wb_dest["Physicians"]

def strip_phone_fax(num_as_string):
    temp_strip = re.sub(r"[\() -]+","",num_as_string)
    final_number = temp_strip[:3]+"."+temp_strip[3:6]+"."+temp_strip[6:]
    return final_number


def create_main_org(row):
    for cell in row:
        if cell.value == "Facility":
            cell.value = "Main Org"
        else:
            continue
        return row


def prep_spreadsheet(a_range):
    for row in ws_source.iter_rows(a_range):
        if row[0].value:
            create_main_org(row)
        elif not row[1].value:
            previous_row_id = str(int(row[1].row) - 1)
            ws_source[str(row[1].column) + str(row[1].row)].value = ws_source[str(row[1].column) + previous_row_id].value
        else:
            continue
    wb_source.save(source_file)


def copy_locations(a_range):
    count = 2
    for row in ws_source.iter_rows(a_range):
        if row[13].value == "Main Org":
            ws_dest['A' + str(count)].value = row[1].value
            ws_dest['B' + str(count)].value = row[0].value
            ws_dest['C' + str(count)].value = row[14].value
            ws_dest['D' + str(count)].value = row[14].value
            ws_dest['E' + str(count)].value = row[15].value
            ws_dest['F' + str(count)].value = row[16].value
            ws_dest['H' + str(count)].value = row[17].value
            ws_dest['I' + str(count)].value = row[18].value
            ws_dest['J' + str(count)].value = row[19].value
            try:
                ws_dest['K' + str(count)].value = strip_phone_fax(row[20].value)
                ws_dest['L' + str(count)].value = strip_phone_fax(row[22].value)
            except TypeError as e:
                pass
        elif row[13].value == "Facility":
            ws_dest['A' + str(count)].value = row[1].value
            ws_dest['C' + str(count)].value = row[14].value
            ws_dest['D' + str(count)].value = row[14].value
            ws_dest['E' + str(count)].value = row[15].value
            ws_dest['F' + str(count)].value = row[16].value
            ws_dest['H' + str(count)].value = row[17].value
            ws_dest['I' + str(count)].value = row[18].value
            ws_dest['J' + str(count)].value = row[19].value
            try:
                ws_dest['K' + str(count)].value = strip_phone_fax(row[20].value)
                ws_dest['L' + str(count)].value = strip_phone_fax(row[22].value)
            except TypeError as e:
                pass
        else:
            continue
        count += 1
    wb_dest.save(dest_file)


def copy_physicians(a_range):
    count = 2
    for row in ws_source.iter_rows(a_range):
        if row[13].value == "Doctor":
            #print(re.findall(r"[\w]+",row[14].internal_value))
            #print(strip_phys_name(row[14].internal_value))
            wsPhys_dest['A' + str(count)].value = row[1].value
            wsPhys_dest['B' + str(count)].value, wsPhys_dest['C' + str(count)].value = strip_phys_name(row[14].value)
            wsPhys_dest['D' + str(count)].value = row[23].value
            try:
                wsPhys_dest['E' + str(count)].value = strip_phone_fax(row[20].value)
                wsPhys_dest['F' + str(count)].value = strip_phone_fax(row[22].value)
            except TypeError as e:
                pass
        else:
            continue
        count += 1
    wb_dest.save(dest_file)


def strip_phys_name(name_as_string):
    #split_name_list = re.findall(r"[\w]+",name_as_string)
    split_name_list = name_as_string.split()
    first_name = split_name_list[1]
    last_name = split_name_list[0]
    return first_name, last_name


prep_spreadsheet(total_range)
copy_locations(total_range)
copy_physicians(total_range)
