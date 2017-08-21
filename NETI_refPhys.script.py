"""
quick script to translate the provided XML and move to excel
"""
import xml.etree.ElementTree as ET
import re
import json
import requests

from openpyxl import load_workbook

# add xml to variable
xml_phys = "NETIPhysiciansNoLimit.xml"
xml_icd = "NETI_ICDsince2015.xml"
# add excels to variable
xls_phys = "NETI_refPhys2.xlsx"
xls_icd = "NETI_ICD.xlsx"
# define range for ref phys
phys_range = "B3:F5000"
icd_range = "B2:B700"
icd_desc_range = "B3:C405"
# loading workbooks
wb = load_workbook(xls_phys)
wb_icd = load_workbook(xls_icd)


def convert_phys(xml_doc):
    tree = ET.parse(xml_doc)
    root = tree.getroot()
    lst = []
    for row in root[4][0]:
        name_lst = []
        for cell in row:
            for data in cell:
                try:
                    if any(char.isdigit() for char in data.text):
                        number_stripped = re.sub('[^0-9]', "", str(data.text))
                        if len(number_stripped) > 10:
                            number_stripped = number_stripped[1:]
                            name_lst.append(number_stripped)
                        else:
                            name_lst.append(number_stripped)
                    else:
                        name_stripped = str(data.text).strip()
                        # print('Stripped name {}'.format(name_stripped))
                        name_split = name_stripped.split()
                        # print('Split name {}'.format(name_split))
                        if len(name_split) == 3:
                            item_name_dict = {'first_name': name_split[0],
                                              'last_name': name_split[1],
                                              'qualification': name_split[2],
                                              }
                            name_lst.append(item_name_dict)
                        elif len(name_split) == 4:
                            item_name_dict = {'first_name': name_split[0],
                                              'middle_name': name_split[1],
                                              'last_name': name_split[2],
                                              'qualification': name_split[3],
                                              }
                            name_lst.append(item_name_dict)
                        else:
                            print('There was an issue parsing the name for {}'.format(data.text))
                            name_lst.append(data.text)
                except TypeError:
                    name_lst.append(data.text)
        lst.append(name_lst)
    return lst


def add_to_excel(lst, wb, rnge):
    ws = wb.active
    for row in ws.iter_rows(rnge):
        for item in lst[2:]:
            try:
                row[0].value = item[0]['first_name']
                try:
                    row[1].value = item[0]['middle_name']
                except KeyError:
                    pass
                row[2].value = item[0]['last_name']
                row[3].value = item[0]['qualification']
            except TypeError:
                row[3].value = item[0]
            row[4].value = item[1]
            lst.remove(item)
            break

    wb.save(xls_phys)
# comment out, script already executed 7/6/17
# add_to_excel(convert_phys(xml_phys), wb, phys_range)


def convert_icd(xml_doc, wb, a_range):
    # getting trunk of xml
    tree = ET.parse(xml_doc)
    root = tree.getroot()
    lst = []
    ws = wb.active
    for row in root[4][0]:
        for cell in row:
            for data in cell:
                lst.append(data.text)
    for row in ws.iter_rows(a_range):
        for item in lst:
            row[0].value = item
            lst.remove(item)
            break

    wb.save(xls_icd)
# comment out, script already executed on 7/6/17
# convert_icd(xml_icd, wb_icd, icd_range)


def add_icd_descript(wb, a_range):
    ws = wb.active
    url = 'http://icd10api.com/?'
    query_str = 'code={}&desc=short&r=json'
    count = 3
    for row in ws.iter_rows(a_range):
        final_str = url+query_str.format(row[0].value)
        print("{} corresponds to row {}".format(final_str, count))
        try:
            response = requests.get(final_str)
        except (ValueError, KeyError):
            row[1].value = "Error"
        res_json = json.loads(response.text)
        row[1].value = res_json['Description']
        count += 1

    wb.save(xls_icd)

add_icd_descript(wb_icd, icd_desc_range)
