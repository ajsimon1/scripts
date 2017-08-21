
"""

author Adam Simon for VitalAxis
1/24/2017

"""
from openpyxl import *
from collections import OrderedDict


of_source_file = r"C:\Users\adam\Documents\VitalAxis\Sagis Dx\Sagis_OF.OP_final.xlsx"
lk_of_file = r'C:\Users\adam\Documents\VitalAxis\Sagis Dx\Kellison_Client Configuration-PV1-13.xlsx'
txt_file = r'C:\Users\Adam\Documents\VitalAxis\Sagis Dx\Sagis_lk.va_diffs.txt'
thej_guid_file = r'C:\Users\Adam\Documents\VitalAxis\Sagis Dx\Sagis_Facilities.GUID_thejesh.xlsx'

va_range = "A2:D180"
lk_range = "A2:C207"
thej_range = "A2:E183"


va_wb = load_workbook(of_source_file)
lk_wb = load_workbook(lk_of_file)

# load active sheetnames
va_ws = va_wb['Facilities']
lk_ws = lk_wb.active
# create dctionaries
lk_dict = {}
va_dict = {}


def create_dict(a_sheet, a_range, a_dict):
    for row in a_sheet.iter_rows(a_range):
        try:
            a_dict[int(row[0].value)] = str(row[1].value)
            ordered_dict = OrderedDict(sorted(a_dict.items(), key=lambda t: t[0]))
        except TypeError:
            a_dict['NULL'] = str(row[1].value)
    return ordered_dict


create_dict(lk_ws, lk_range, lk_dict)
create_dict(va_ws, va_range, va_dict)
# va_count = Counter(va_dict)
# lk_count = Counter(lk_dict)


def get_diffs_in_dics(dict_1, dict_2, file_name):
    doc = open(file_name, 'w')
    dict_1_set = set(dict_2.keys())
    dict_2_set = set(dict_1.keys())
    in_dict_2_not_in_dict_1 = list(dict_2_set.difference(dict_1_set))
    in_dict_1_not_in_dict_2 = list(dict_1_set.difference(dict_2_set))
    dict_set_diff_list = [in_dict_2_not_in_dict_1, in_dict_1_not_in_dict_2]
    for dict_diff_list in dict_set_diff_list:
        print('This is the start of {}'.format(dict_diff_list))
        for item in dict_diff_list:
            try:
                doc.write(dict_diff_list[int(item)])
                doc.close()
            except ValueError:
                continue
            
def add_clientcode_to_thejesh_list(wb1_filename, range1, wb2_filename, range2):
    wb1 = load_workbook(wb1_filename)
    wb2 = load_workbook(wb2_filename)
    ws1 = wb1.active
    ws2 = wb2.active
    for row in ws1.iter_rows(range1):
        for snd_row in ws2.iter_rows(range2):
            if str(row[2].value) == str(snd_row[0].value):
                snd_row[4].value = row[0].value
    wb1.save(wb1_filename)
    wb2.save(wb2_filename)
    
def add_guid_to_lk(wb1_filename, range1, wb2_filename, range2):
    wb1 = load_workbook(wb1_filename)
    wb2 = load_workbook(wb2_filename)
    ws1 = wb1.active
    ws2 = wb2.active
    for row in ws1.iter_rows(range1):
        for snd_row in ws2.iter_rows(range2):
            print('comparing {} with {}'.format(str(row[0].value), str(snd_row[4].value)))
            if str(row[0].value) == str(snd_row[4].value):
                row[2].value = snd_row[2].value
                break
            else:
                row[2].value = 'NEEDED'
                
    wb1.save(wb1_filename)
    wb2.save(wb2_filename)
    
def add_org_to_thejesh_list(wb1_filename, range1, wb2_filename, range2):
    wb1 = load_workbook(wb1_filename)
    wb2 = load_workbook(wb2_filename)
    ws1 = wb1.active
    ws2 = wb2.active
    for row in ws1.iter_rows(range1):
        for snd_row in ws2.iter_rows(range2):
            # print('comparing {} with {}'.format(str(row[1].value), str(snd_row[2].value)))
            if str(row[1].value) == str(snd_row[2].value):
                row[0].value = snd_row[1].value
                break
            else:
                continue
    wb1.save(wb1_filename)
    wb2.save(wb2_filename)
                       
add_org_to_thejesh_list(thej_guid_file, thej_range, of_source_file, va_range)
# add_guid_to_lk(lk_of_file, lk_range, thej_guid_file, thej_range)
# add_clientcode_to_thejesh_list(of_source_file, 'A2:D180', thej_guid_file, 'B3:F178')
# get_diffs_in_dics(va_dict, lk_dict, txt_file)
# print(va_dict)
